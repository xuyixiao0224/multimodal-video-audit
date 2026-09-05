"""
video_excel_utils.py
视频分析Excel工具模块 (5列独立版)

功能特点：
1. 另存为新文件：自动生成 "原文件名_视频文件夹名.xlsx"
2. 模糊匹配文件名：自动忽略 (1), (2) 等下载后缀
3. 关键特性：将错误详情拆分为 5 列 (Start, End, Category, Reason, Content)
4. 样式美化：自动合并左侧相同信息的单元格 (A-J列)
"""

import json
import os
import re
import pandas as pd
from typing import Dict, Optional
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def normalize_filename(filename: str) -> str:
    """
    清洗文件名用于匹配
    将 "video_name (1).mp4" 清洗为 "video_name.mp4"
    """
    cleaned = re.sub(r'\s*\(\d+\)(?=\.)', '', filename)
    return cleaned

def write_video_results_to_excel(excel_path: str, results: Dict[str, Dict], video_folder_name: Optional[str] = None):
    """
    将视频分析结果写入新Excel文件 (5列独立版)
    """
    try:
        # 1. 构造新文件路径
        dir_name = os.path.dirname(excel_path)
        base_name = os.path.basename(excel_path)
        file_name_no_ext, ext = os.path.splitext(base_name)
        
        suffix = f"_{video_folder_name}" if video_folder_name else "_analyzed"
        new_filename = f"{file_name_no_ext}{suffix}{ext}"
        output_path = os.path.join(dir_name, new_filename)
        
        print(f"[System] 正在读取原始 Excel: {excel_path}")
        original_df = pd.read_excel(excel_path)

        # 2. 自动检测包含视频URL的列
        video_column_idx = None
        video_col_name = None
        
        for col_idx, col_name in enumerate(original_df.columns):
            sample_values = original_df.iloc[:10, col_idx].astype(str).values
            for val in sample_values:
                if '.mp4' in val:
                    video_column_idx = col_idx
                    video_col_name = col_name
                    break
            if video_column_idx is not None:
                break

        if video_column_idx is None:
            print("❌ 警告：未找到包含视频URL的列")
            return

        print(f"[Info] 视频URL列检测为: '{video_col_name}' (第 {video_column_idx+1} 列)")

        # 3. 构建新的数据行
        new_rows = []
        
        # 预处理 results
        normalized_results = {}
        for file_path, result in results.items():
            raw_filename = os.path.basename(file_path)
            clean_name = normalize_filename(raw_filename)
            normalized_results[clean_name] = result

        processed_videos = 0

        for idx, row in original_df.iterrows():
            video_url = str(row.iloc[video_column_idx])
            url_filename = video_url.split('/')[-1] if '/' in video_url else video_url
            url_filename_clean = normalize_filename(url_filename)
            
            matched_result = None
            
            # 匹配逻辑
            if url_filename_clean in normalized_results:
                matched_result = normalized_results[url_filename_clean]
            else:
                for res_name, res_data in normalized_results.items():
                    if res_name in url_filename_clean or url_filename_clean in res_name:
                        matched_result = res_data
                        break
            
            base_row_data = row.to_dict()

            if matched_result:
                processed_videos += 1
                errors = matched_result.get("all_detected_errors", [])
                
                if errors:
                    # == 场景A: 有错误，拆分为多行，填入5列 ==
                    for err in errors:
                        new_row = base_row_data.copy()
                        
                        # 1. Start Time
                        new_row['AI_Start_Time'] = err.get('timestamp_start', '')
                        
                        # 2. End Time
                        new_row['AI_End_Time'] = err.get('timestamp_end', '')
                        
                        # 3. Category (对应 code 或 category)
                        # 优先取 category，没有则取 code
                        new_row['AI_Category'] = err.get('category', err.get('code', ''))
                        
                        # 4. Reason
                        new_row['AI_Reason'] = err.get('reason', '')
                        
                        # 5. Content (场景描述)
                        new_row['AI_Content'] = err.get('content', '') 
                        
                        new_rows.append(new_row)
                else:
                    # == 场景B: 无错误 (PASS) ==
                    new_row = base_row_data.copy()
                    new_row['AI_Start_Time'] = ''
                    new_row['AI_End_Time'] = ''
                    new_row['AI_Category'] = 'PASS'
                    new_row['AI_Reason'] = '未发现明显错误'
                    new_row['AI_Content'] = ''
                    new_rows.append(new_row)
            else:
                # == 场景C: 未分析 ==
                new_row = base_row_data.copy()
                # 补全空列
                for col in ['AI_Start_Time', 'AI_End_Time', 'AI_Category', 'AI_Reason', 'AI_Content']:
                    new_row[col] = ''
                new_rows.append(new_row)

        # 4. 保存新文件
        new_df = pd.DataFrame(new_rows)
        new_df.to_excel(output_path, index=False, engine='openpyxl')
        
        # 5. 美化表格 (合并 A-J 列)
        print(f"[System] 正在美化表格格式 (合并 A-J 列)...")
        wb = load_workbook(output_path)
        ws = wb.active
        
        MERGE_END_COL = 10  # 假设前10列是元数据
        
        new_video_col_idx = None
        for i, cell in enumerate(ws[1]):
            if cell.value == video_col_name:
                new_video_col_idx = i + 1
                break
        
        if new_video_col_idx:
            current_video_val = None
            merge_start_row = 2
            
            for row in range(2, ws.max_row + 2):
                if row <= ws.max_row:
                    cell_val = ws.cell(row=row, column=new_video_col_idx).value
                else:
                    cell_val = "END_OF_FILE"

                if cell_val != current_video_val:
                    if current_video_val is not None and (row - merge_start_row) > 1:
                        for col in range(1, MERGE_END_COL + 1):
                            ws.merge_cells(start_row=merge_start_row, start_column=col, 
                                         end_row=row-1, end_column=col)
                            cell = ws.cell(row=merge_start_row, column=col)
                            cell.alignment = Alignment(vertical='center', horizontal='left')

                    current_video_val = cell_val
                    merge_start_row = row

        wb.save(output_path)
        print(f"✅ 新报告已生成: {output_path}")
        print(f"   已新增5列: AI_Start_Time, AI_End_Time, AI_Category, AI_Reason, AI_Content")

    except Exception as e:
        print(f"❌ 写入Excel失败: {e}")
        import traceback
        traceback.print_exc()